from prettytable import PrettyTable
from bleak import BleakScanner,BleakClient
from color import *
import sys,getopt,asyncio,os,readline
from pyfiglet import Figlet
from openpyxl import load_workbook

##################################################################
#处理表格,检索厂商信息
file = "./CompanyIdentfiers.xlsx"   #filepath

class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def get_rows(self):
        rows = self.ws.max_row
        return rows

    def get_clos(self):
        clo = self.ws.max_column
        return clo

    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

excel_utils = ExcelUtils()
excel_dict = {}
row = excel_utils.get_rows()
for i in range(1,row+1):
    dict_key = excel_utils.get_cell_value(i, 1)
    dict_value = excel_utils.get_cell_value(i, 2)
    excel_dict[dict_key] = dict_value

##################################################################

async def connclient(address,device):
    client = BleakClient(address,adapter=device)
    await client.connect()
    return client

async def disconnclient(client):
    await client.disconnect()

async def scan_devices(device):
    devices = await BleakScanner.discover(adapter=device)
    BLEdevices = PrettyTable([blue("编号"), yellow("设备地址"), green("设备名"), green("RSSI"), green("类型"), green("厂商")])
    BLEdevices.align[green("厂商")] = "l"
    for i in range(0,len(devices)):
        try:
            ManufacturerData = devices[i].details["props"]["ManufacturerData"].keys()
            company = list(ManufacturerData)[0]
            BLEdevices.add_row([blue(str(i)), yellow(devices[i].address), green(devices[i].name),  green(str(devices[i].rssi)), green(str(devices[i].details["props"]["AddressType"])), green(str(excel_dict[company]))])
        except:
            BLEdevices.add_row([blue(str(i)), yellow(devices[i].address), green(devices[i].name),  green(str(devices[i].rssi)), green(str(devices[i].details["props"]["AddressType"])), green("Unknow")])
    print(BLEdevices)

async def scan_services(client):
    SvcS = PrettyTable([blue("服务UUID"), yellow("句柄")])
    svcs = await client.get_services()
    for service in svcs:
        SvcS.add_row([blue(service.uuid),yellow(str(hex(service.handle)))])
    print(SvcS)

async def scan_characteristics(client,serviceid):
    CharS = PrettyTable([blue("特性UUID"), yellow("句柄"), green("属性")])
    svcs = await client.get_services()
    for service in svcs:
        if service.uuid == serviceid:
            for char in service.characteristics:
                CharS.add_row([blue(char.uuid),yellow(str(hex(int(char.handle)+1))),green('; '.join(char.properties))])
        else:
            pass
    CharS.align[green("属性")] = 'l'
    CharS.sortby = yellow("句柄")
    print(CharS)

async def read_value(client,char_uuid):
    try:
        value = await client.read_gatt_char(char_uuid)
        print(green("[+]RECV: ") + str(value))
    except:
        print(red("[x]ERROR: Can't read value from ") + str(char_uuid))

async def write_value(client,char_uuid,string):
    if string[:4] == "hex:":
        value = string[4:]
        try:
            value = bytearray.fromhex(value)
            await client.write_gatt_char(char_uuid,value)
            print(green("[+]SEND RAW SUCCESS!"))
        except:
            print(red("[x]ERROR: Can't write value to ") + str(char_uuid))
    else:
        value = string
        try:
            await client.write_gatt_char(char_uuid,value.encode())
            print(green("[+]SEND STR SUCCESS!"))
        except:
            print(red("[x]ERROR: Can't write value to ") + str(char_uuid))

def callback(sender: int, data: bytearray):
    print(green("[+]RECV: ") + f"{data}")

async def listen_notify(client,char_uuid,time,value):
    await client.start_notify(char_uuid, callback)
    await write_value(client,char_uuid,value)
    await asyncio.sleep(time)
    await client.stop_notify(char_uuid)

def show_Adapters():
    ap_address = os.popen("hciconfig | grep 'BD Address' | awk '{print $3}'").read().replace('\n', ';')[:-1].split(";")
    device = os.popen("hciconfig | grep hci | awk '{print $1}'").read().replace(':\n', ';')[:-1].split(";")
    Adapter = PrettyTable(["适配器", "地址"])
    for i in range(0,len(ap_address)):
        Adapter.add_row([yellow(device[i]),blue(ap_address[i])])
    print(Adapter)

###############################################################

async def main():
    f = Figlet(font="slant", width=100)
    print(f.renderText("BLE Security Tool"))
    print(" Author: yichen               Version: 1.02\n")
    meun = PrettyTable(["选项", "说明"])
    meun.align["选项"] = 'l'
    meun.align["说明"] = 'l'
    meun.add_row([yellow("help / h"),blue("展示帮助菜单")])
    meun.add_row([yellow("lescan"),blue("扫描周围低功耗蓝牙设备")])
    meun.add_row([yellow("hciconfig"),blue("展示蓝牙适配器")])
    meun.add_row([yellow("connect"),blue("连接BLE设备")])
    meun.add_row([yellow("disconnect"),blue("断开BLE设备的连接")])
    meun.add_row([yellow("services"),blue("扫描已连接设备的所有服务")])
    meun.add_row([yellow("characteristics"),blue("扫描某一服务的所有特性")])
    meun.add_row([yellow("read"),blue("读取某一特性的值")])
    meun.add_row([yellow("write"),blue("向某一特性写值")])
    meun.add_row([yellow("listen"),blue("监听某特性值的返回值")])
    meun.add_row([yellow("readddd"),blue("指定次数读取特性的值")])
    meun.add_row([yellow("force"),blue("发送指定范围的值给某特性")])
    meun.add_row([yellow("bdaddr"),blue("修改MAC地址")])
    meun.add_row([yellow("restart"),blue("重启蓝牙服务")])
    print(meun)
    choose = ""
    while choose != "exit":
        choose = input("--> ")
        if choose == "lescan":
            try:
                await scan_devices("hci0")
            except:
                show_Adapters()
                adapter = input("Adapter: ")
                await scan_devices(adapter)
        if choose == "hciconfig":
            show_Adapters()
        if choose == "connect":
            address = input("MAC Address: ")
            try:
                client = await connclient(address,"hci0")
            except:
                show_Adapters()
                adapter = input("Adapter: ")
                client = await connclient(address,adapter)
        if choose == "disconnect":
            await disconnclient(client)
        if choose == "services":
            await scan_services(client)
        if choose == "characteristics":
            serviceid = input("service uuid: ")
            await scan_characteristics(client,serviceid)
        if choose == "read":
            char_uuid = input("characteristics uuid: ")
            await read_value(client,char_uuid)
        if choose == "write":
            char_uuid = input("characteristics uuid: ")
            string = input("input: ")
            await write_value(client,char_uuid,string)
        if choose == "listen":
            char_uuid = input("characteristics uuid: ")
            time = input("listen time(default 3): ") or 3
            value = input("input(default 'hello'): ") or "hello"
            await listen_notify(client,char_uuid,int(time),value)
        if choose == "bdaddr":
            print(red("需要在本目录make一下bdaddr"))
            old_addr = input("now addr: ")
            new_addr = input("new addr: ")
            command = "./bdaddr -i " + old_addr + " " + new_addr
            result = os.popen(command).read()
            print(result)
        if choose == "readddd":
            time_num = input("number of times: ")
            char_uuid = input("characteristics uuid: ")
            for i in range(0,int(time_num)):
                await read_value(client,char_uuid)
        if choose == "force":
            start_num = int(input("start num: "))
            stop_num = int(input("stop num: "))
            char_uuid = input("characteristics uuid: ")
            for i in range(start_num,stop_num):
                x = str(hex(i))[2:]
                x = x.zfill(2)
                string = "hex:" + x
                await write_value(client,char_uuid,string)
        if choose == "clear":
            sys.stdout.write("\x1b[2J\x1b[H")
        if choose == "restart":
            os.system("service bluetooth restart")
        if choose == "h" or choose == "help":
            print(meun)

try:
    asyncio.run(main())
except:
    asyncio.get_event_loop().run_until_complete(main())